import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("age.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]

print(sheet["A1"].value)
print(sheet.cell(1, 1).value)

print(sheet.max_row)  # Number of rows

sheet.cell(1,4).value = "Age +10"
for row in range(2, sheet.max_row+1):
    age_after_10_cell = sheet.cell(row, 4)
    age_after_10_cell.value = sheet.cell(row, 3).value + 10

wb.save("age_after_10.xlsx")

wb1 = xl.load_workbook("age_after_10.xlsx")
sheet1 =wb1["Sheet1"]

print(sheet1.cell(2,4).value)
#values = Reference(sheet1,min_row=2,max_row=sheet1.max_row, min_col=4,max_col=4)
values = Reference(sheet1, min_col=4, min_row=1, max_row=sheet1.max_row, max_col=4)
#friends = Reference(sheet1,min_row=2,max_row=sheet1.max_row, min_col=1,max_col=1)
friends = Reference(sheet1,min_col=1, max_col=1, min_row=2, max_row=sheet1.max_row)

chart = BarChart()
chart.type = "col"
chart.title = "Age Graph"
chart.y_axis.title = 'Age in Years'
chart.x_axis.title = 'Friends'

chart.add_data(values,titles_from_data=True)

chart.set_categories(friends)
sheet1.add_chart(chart,"E2")

wb1.save("age_after_10.xlsx")